from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from pathlib import Path


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "students.xlsx"

workbook: Workbook = load_workbook(WORKBOOK_PATH)
worksheet: Worksheet = workbook.active

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)

worksheet['B2'].value = 10.0
workbook.save(WORKBOOK_PATH)
# workbook.save(WORKBOOK_PATH)
