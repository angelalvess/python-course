from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "students.xlsx"

workbook = Workbook()
worksheet: Worksheet = workbook.active

worksheet.cell(row=1, column=1, value="Name")
worksheet.cell(row=1, column=2, value="Grade")
worksheet.cell(row=1, column=3, value="Age")

students = [

    ["John", 9.5, 25],
    ["Jane", 8.5, 22],
    ["Jim", 7.5, 20],
    ["Jill", 6.5, 18],

]

# for i, student_row in enumerate(students, start=2):
#     for j, student_cell in enumerate(student_row, start=1):
#         worksheet.cell(row=i, column=j, value=student_cell)


# Remover uma planilha
# workbook.remove(worksheet)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
